import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select

file = '/Users/vivekkumar/Downloads/simple_interest_data.xlsx'
workbook = openpyxl.load_workbook(file)
sheet = workbook.active

chrome_options = Options()
driver = webdriver.Chrome(options=chrome_options)
driver.get(
    "https://www.moneycontrol.com/fixed-income/calculator/state-bank-of-india-sbi/fixed-deposit-calculator-SBI-BSB001.html")

wait = WebDriverWait(driver, 10)

for row in range(2, sheet.max_row + 1):
    deposit_amount = sheet.cell(row, 1).value
    interest_rate = sheet.cell(row, 2).value
    tenure_value = sheet.cell(row, 3).value
    tenure_type = sheet.cell(row, 4).value.lower()  # Assuming the 4th column specifies 'year(s)', 'month(s)', 'day(s)'
    frequency = sheet.cell(row, 5).value  # Assuming the 5th column specifies the frequency

    deposit_field = wait.until(EC.presence_of_element_located((By.XPATH, "//*[@id='principal']")))
    deposit_field.clear()
    deposit_field.send_keys(str(deposit_amount))

    time.sleep(2)
    rate_field = driver.find_element(By.ID, "interest")
    rate_field.clear()
    rate_field.send_keys(str(interest_rate))

    time.sleep(2)
    tenure_field = driver.find_element(By.ID, "tenure")
    tenure_field.clear()
    tenure_field.send_keys(str(tenure_value))

    # Handle the dropdowns for day, month, and year
    tenure_type_dropdown = Select(driver.find_element(By.ID, "tenurePeriod"))
    if "year" in tenure_type:
        tenure_type_dropdown.select_by_visible_text("year(s)")
    elif "month" in tenure_type:
        tenure_type_dropdown.select_by_visible_text("month(s)")
    elif "day" in tenure_type:
        tenure_type_dropdown.select_by_visible_text("day(s)")

    time.sleep(2)
    # Handle the frequency dropdown
    frequency_dropdown = Select(driver.find_element(By.ID, "frequency"))
    frequency_dropdown.select_by_visible_text(frequency)

    time.sleep(2)
    calculate_button = driver.find_element(By.XPATH, "//*[@id='fdMatVal']/div[2]/a[1]/img")
    calculate_button.click()

    time.sleep(3)

driver.quit()