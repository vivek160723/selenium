# from openpyxl import load_workbook
# from openpyxl import Workbook
# from datetime import datetime
#
# # Load the existing workbook
# workbook = load_workbook('/Users/vivekkumar/Downloads/Data.xlsx')
# sheet = workbook['Sheet1']
#
# # Create a new workbook for the output
# new_workbook = Workbook()
# new_sheet = new_workbook.active
# new_sheet.title = "Processed Data"  # Optional: Set the sheet title
#
# num_rows = sheet.max_row
# num_cols = sheet.max_column
#
# # Write the data to the new sheet
# for row in range(2, num_rows + 1):
#     row_data = []
#     for col in range(1, num_cols + 1):
#         cell_value = sheet.cell(row=row, column=col).value
#         if isinstance(cell_value, datetime):
#             cell_value = cell_value.strftime('%Y-%m-%d')  # Format datetime as string
#         row_data.append(cell_value)
#
#     # Write the row data to the new sheet
#     new_sheet.append(row_data)
#
# # Save the new workbook
# new_workbook.save('/Users/vivekkumar/Downloads/Processed_Data.xlsx')
# workbook.close()

import openpyxl

file = '/Users/vivekkumar/Downloads/Data.xlsx'
workbook = openpyxl.load_workbook(file)

sheet = workbook.active

# Iterate over all rows and columns
for r in range(1, 6):
    for c in range(1, 5):
        cell = sheet.cell(r, c)

        # Check if the current cell is part of a merged range
        if cell.coordinate in sheet.merged_cells:
            continue  # Skip the merged cells

        # Assign value to the cell
        cell.value = "❤"

workbook.save(file)