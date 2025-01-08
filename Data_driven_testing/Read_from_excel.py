from openpyxl import load_workbook
from datetime import datetime

workbook = load_workbook('/Users/vivekkumar/Downloads/Data.xlsx')
sheet = workbook['Sheet1']

num_rows = sheet.max_row
num_cols = sheet.max_column

for row in range(2, num_rows + 1):
    row_data = []
    for col in range(1, num_cols + 1):
        cell_value = sheet.cell(row=row, column=col).value
        if isinstance(cell_value, datetime):
            cell_value = cell_value.strftime('%Y-%m-%d')  # Format the datetime as a string
        row_data.append(cell_value)
    print(row_data)

workbook.close()